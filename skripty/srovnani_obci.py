#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Srovnání Prštic s podobnými obcemi (PLAN.md úkol 2.9, ZADANI P-35).

Dvě srovnání:
  (a) OBJEM ROZPOČTU na obyvatele proti obcím okresu Brno-venkov, které mají
      900–1 100 obyvatel (velikostně nejbližší srovnatelná skupina);
  (b) PRŮMĚRNÝ VĚK proti dvaceti geograficky nejbližším obcím a proti okresu.

Zdroje:
  * počty obyvatel a struktura — ČSÚ, cz0643.xlsx (soukromá zóna)
  * průměrný věk podle obcí — ČSÚ, csu-obyv-obce-2025.xlsx (soukromá zóna)
  * IČO obcí — ARES (veřejné API)
  * rozpočty — MONITOR, rozklikávací rozpočet (konsolidovaná skutečnost)
  * souřadnice obcí — Wikidata (SPARQL)

Výstup: data/srovnani-obci.json

Zásady:
  * Srovnává se TŘÍLETÝ PRŮMĚR 2022–2024, ne jeden rok — jednotlivé roky
    kolísají podle investic a jednorázový výkyv by pořadí zkreslil.
  * Web nepíše „nejhorší/nejlepší obec"; uvádí konkrétní ukazatel a pořadí v něm.
  * Jde o jednorázovou kontextovou rešerši, ne o průběžně aktualizovanou tabulku.

Spuštění: python3 skripty/srovnani_obci.py [--offline]
          (--offline použije uloženou odpověď MONITORu/ARES ze soukromé zóny)
"""
import os, sys, json, time, math, unicodedata, urllib.request
from datetime import datetime

try:
    import openpyxl
except ImportError:
    sys.exit("CHYBA: chybí openpyxl")

PRIV = os.path.expanduser("~/Developer/transparentniprstice-private")
WEB = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
CSU_OBCE = os.path.join(PRIV, "zdroje", "cz0643.xlsx")
CSU_VEK = os.path.join(PRIV, "zdroje", "csu-obyv-obce-2025.xlsx")
CACHE = os.path.join(PRIV, "zdroje", "srovnani-cache.json")

KOD_PRSTICE = "583707"
ROK = 2025                      # poslední rok s počty obyvatel
ROKY_ROZPOCTU = [2022, 2023, 2024]   # uzavřené roky pro tříletý průměr
PASMO = (900, 1100)             # velikostně srovnatelné obce
NEJBLIZSICH = 20

# kontrolní hodnoty — když se rozejdou, něco se ve zdroji změnilo
KONTROLA = {"prstice_obyvatel": 997, "prstice_vek": 43.3, "okres_vek": 41.6}


def http_json(url, telo=None):
    hlavicky = {"User-Agent": "transparentniprstice/1.0", "Accept": "application/json"}
    if telo is not None:
        data = json.dumps(telo).encode()
        hlavicky["Content-Type"] = "application/json"
        req = urllib.request.Request(url, data=data, headers=hlavicky)
    else:
        req = urllib.request.Request(url, headers=hlavicky)
    with urllib.request.urlopen(req, timeout=30) as r:
        return json.load(r)


def bez_diakritiky(s):
    s = unicodedata.normalize("NFKD", str(s)).strip().lower()
    return "".join(c for c in s if not unicodedata.combining(c))


def nacti_obyvatele():
    """kód obce -> (název, počet obyvatel k 1.1.ROK)"""
    ws = openpyxl.load_workbook(CSU_OBCE, read_only=True, data_only=True)["CZ0643"]
    out = {}
    for r in ws.iter_rows(values_only=True):
        if r[0] == "Rok" or r[0] is None:
            continue
        try:
            rok = int(r[0])
        except (TypeError, ValueError):
            continue
        if rok == ROK and r[4] not in (None, ""):
            out[str(r[1])] = (r[2], int(r[4]))
    return out


def nacti_vek():
    """kód obce -> průměrný věk (celá ČR, filtruje se podle okresu)"""
    ws = openpyxl.load_workbook(CSU_VEK, read_only=True, data_only=True)["List1"]
    out = {}
    for r in ws.iter_rows(values_only=True):
        kod = r[1]
        if kod and str(kod).isdigit() and len(str(kod)) == 6 and r[6] not in (None, ""):
            try:
                out[str(kod)] = float(r[6])
            except (TypeError, ValueError):
                pass
    return out


def ares_ico(kod_obce):
    d = http_json("https://ares.gov.cz/ekonomicke-subjekty-v-be/rest/ekonomicke-subjekty/vyhledat",
                  {"sidlo": {"kodObce": int(kod_obce)}, "pravniForma": ["801"], "pocet": 5, "start": 0})
    for s in d.get("ekonomickeSubjekty", []):
        if s.get("sidlo", {}).get("kodObce") == int(kod_obce):
            return s["ico"]
    subs = d.get("ekonomickeSubjekty", [])
    return subs[0]["ico"] if subs else None


def monitor_rozpocet(ico, rok):
    """Konsolidovaná skutečnost (příjmy, výdaje) z rozklikávacího rozpočtu."""
    url = (f"https://monitor.statnipokladna.gov.cz/api/rozpocet/souhrnny"
           f"?obdobi={rok - 2000:02d}12&ic={ico}")
    try:
        d = http_json(url)
    except Exception:
        return None, None
    prijmy = vydaje = None
    for ch in d.get("children", []):
        hodnota = (ch.get("budget") or {}).get("reality")
        if ch.get("name") == "Revenues":
            prijmy = hodnota
        elif ch.get("name") == "Expenditures":
            vydaje = hodnota
    return prijmy, vydaje


def souradnice_wikidata():
    """název obce (bez diakritiky) -> (lat, lon) pro okres Brno-venkov"""
    dotaz = """SELECT ?obecLabel ?lat ?lon WHERE {
      ?d rdfs:label "okres Brno-venkov"@cs . ?obec wdt:P131 ?d ; wdt:P625 ?c .
      BIND(geof:latitude(?c) AS ?lat) BIND(geof:longitude(?c) AS ?lon)
      SERVICE wikibase:label { bd:serviceParam wikibase:language "cs". } }"""
    url = "https://query.wikidata.org/sparql?format=json&query=" + urllib.parse.quote(dotaz)
    d = http_json(url)
    out = {}
    for r in d["results"]["bindings"]:
        out.setdefault(bez_diakritiky(r["obecLabel"]["value"]),
                       (float(r["lat"]["value"]), float(r["lon"]["value"])))
    return out


def vzdalenost(a, b):
    R = 6371.0
    la1, lo1, la2, lo2 = map(math.radians, [a[0], a[1], b[0], b[1]])
    h = math.sin((la2 - la1) / 2) ** 2 + math.cos(la1) * math.cos(la2) * math.sin((lo2 - lo1) / 2) ** 2
    return 2 * R * math.asin(math.sqrt(h))


def main():
    offline = "--offline" in sys.argv
    for p in (CSU_OBCE, CSU_VEK):
        if not os.path.exists(p):
            sys.exit(f"CHYBA: chybí zdroj {p}")

    obyv, vek = nacti_obyvatele(), nacti_vek()
    vek_okres = {k: v for k, v in vek.items() if k in obyv}

    # kontrola proti známým hodnotám
    if obyv[KOD_PRSTICE][1] != KONTROLA["prstice_obyvatel"]:
        sys.exit(f"CHYBA: Prštice mají {obyv[KOD_PRSTICE][1]} obyv., čekáno {KONTROLA['prstice_obyvatel']}")
    if abs(round(vek_okres[KOD_PRSTICE], 1) - KONTROLA["prstice_vek"]) > 0.05:
        sys.exit("CHYBA: průměrný věk Prštic nesedí na kontrolu")

    cache = json.load(open(CACHE, encoding="utf-8")) if os.path.exists(CACHE) else {}

    # ── (a) rozpočty srovnatelných obcí ──────────────────────────────────────
    skupina = sorted([(v[1], k, v[0]) for k, v in obyv.items() if PASMO[0] <= v[1] <= PASMO[1]])
    rozpocty = []
    for pocet, kod, nazev in skupina:
        zaznam = cache.get("obce", {}).get(kod)
        if zaznam is None:
            if offline:
                continue
            ico = ares_ico(kod); time.sleep(0.2)
            roky = {}
            for r in ROKY_ROZPOCTU:
                p, v = monitor_rozpocet(ico, r); time.sleep(0.15)
                if p is not None:
                    roky[str(r)] = {"prijmy": round(p), "vydaje": round(v or 0)}
            zaznam = {"ico": ico, "roky": roky}
            cache.setdefault("obce", {})[kod] = zaznam
            print(f"    staženo: {nazev}")
        prijmy = [zaznam["roky"][str(r)]["prijmy"] for r in ROKY_ROZPOCTU if str(r) in zaznam["roky"]]
        lidi = [obyv[kod][1]]
        if not prijmy:
            continue
        prumer = sum(prijmy) / len(prijmy)
        rozpocty.append({"kod": kod, "obec": nazev, "ico": zaznam["ico"], "obyvatel": pocet,
                         "prumer_prijmu_kc": round(prumer),
                         "na_obyvatele_kc": round(prumer / (sum(lidi) / len(lidi)))})
    rozpocty.sort(key=lambda x: x["na_obyvatele_kc"])
    poradi = [x["kod"] for x in rozpocty].index(KOD_PRSTICE) + 1
    hodnoty = [x["na_obyvatele_kc"] for x in rozpocty]
    median = (hodnoty[len(hodnoty) // 2] if len(hodnoty) % 2
              else round((hodnoty[len(hodnoty) // 2 - 1] + hodnoty[len(hodnoty) // 2]) / 2))

    # ── (b) průměrný věk ────────────────────────────────────────────────────
    prumer_okres = round(sum(vek_okres[k] * obyv[k][1] for k in vek_okres) /
                         sum(obyv[k][1] for k in vek_okres), 2)
    if abs(prumer_okres - KONTROLA["okres_vek"]) > 0.06:
        sys.exit(f"CHYBA: okresní věk {prumer_okres} ≠ kontrola {KONTROLA['okres_vek']}")

    coords = cache.get("souradnice")
    if coords is None and not offline:
        coords = souradnice_wikidata()
        cache["souradnice"] = coords
    coords = {k: tuple(v) for k, v in (coords or {}).items()}
    dom = coords.get(bez_diakritiky(obyv[KOD_PRSTICE][0]))
    blizke = []
    if dom:
        for kod, (nazev, pocet) in obyv.items():
            if kod == KOD_PRSTICE:
                continue
            c = coords.get(bez_diakritiky(nazev))
            if c and kod in vek_okres:
                blizke.append({"obec": nazev, "km": round(vzdalenost(dom, c), 1),
                               "obyvatel": pocet, "vek": round(vek_okres[kod], 1)})
        blizke.sort(key=lambda x: x["km"])
        blizke = blizke[:NEJBLIZSICH]
    veky = sorted(x["vek"] for x in blizke)
    starsich = sum(1 for k in vek_okres if vek_okres[k] > vek_okres[KOD_PRSTICE])

    json.dump(cache, open(CACHE, "w", encoding="utf-8"), ensure_ascii=False)

    vysledek = {
        "meta": {
            "ukol": "P-35 — jednorázová kontextová rešerše, ne průběžně aktualizovaná tabulka",
            "skupina": f"obce okresu Brno-venkov s {PASMO[0]}–{PASMO[1]} obyvateli",
            "metodika_rozpocet": f"tříletý průměr konsolidovaných příjmů {ROKY_ROZPOCTU[0]}–{ROKY_ROZPOCTU[-1]} "
                                 f"dělený počtem obyvatel k 1. 1. {ROK}",
            "zdroje": ["ČSÚ — počty obyvatel a průměrný věk podle obcí",
                       "ARES — IČO obcí", "MONITOR MF ČR — rozklikávací rozpočet",
                       "Wikidata — souřadnice obcí"],
            "vygenerovano": datetime.now().strftime("%Y-%m-%d %H:%M")},
        "rozpocet": {"obci_ve_skupine": len(rozpocty), "poradi_prstic_od_nejnizsiho": poradi,
                     "prstice_na_obyvatele_kc": next(x["na_obyvatele_kc"] for x in rozpocty if x["kod"] == KOD_PRSTICE),
                     "median_kc": median, "obce": rozpocty},
        "vek": {"prstice": round(vek_okres[KOD_PRSTICE], 1), "okres": prumer_okres,
                "obci_v_okrese": len(vek_okres),
                "starsich_obci_v_okrese": starsich,
                "poradi_prstic_od_nejstarsiho": starsich + 1,
                "nejblizsich": len(blizke),
                "prumer_nejblizsich": round(sum(veky) / len(veky), 1) if veky else None,
                "median_nejblizsich": veky[len(veky) // 2] if veky else None,
                "max_km": blizke[-1]["km"] if blizke else None,
                "obce": blizke}}

    with open(os.path.join(WEB, "data", "srovnani-obci.json"), "w", encoding="utf-8") as f:
        json.dump(vysledek, f, ensure_ascii=False, indent=1)

    r, v = vysledek["rozpocet"], vysledek["vek"]
    print(f"HOTOVO — data/srovnani-obci.json")
    print(f"  rozpočet: Prštice {r['prstice_na_obyvatele_kc']:,} Kč/obyv. → "
          f"{r['poradi_prstic_od_nejnizsiho']}. nejnižší z {r['obci_ve_skupine']} "
          f"(medián {r['median_kc']:,} Kč)".replace(",", " "))
    print(f"  věk:      Prštice {v['prstice']} | okolí {v['prumer_nejblizsich']} | "
          f"okres {v['okres']} → {v['poradi_prstic_od_nejstarsiho']}. nejstarší z {v['obci_v_okrese']}")
    print("  kontroly PASS ✓")


if __name__ == "__main__":
    import urllib.parse
    main()
